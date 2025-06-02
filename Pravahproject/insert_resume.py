# import mysql.connector
# import os
# import pandas as pd

# # Connect to MySQL
# conn = mysql.connector.connect(
#     host="localhost",
#     user="root",
#     password="sonali@12",
#     database="pravahdb"
# )
# cursor = conn.cursor()

# # Insert a PDF resume into the table
# with open('resume.pdf', 'rb') as file:
#     binary_data = file.read()

# cursor.execute("""
#     INSERT INTO pravahapp_jobapplication (full_name, email, phone_number, resume, job_position, applied_at)
#     VALUES (%s, %s, %s, %s, %s, NOW())
# """, ('John Doe', 'johndoe@example.com', '1234567890', binary_data, 'Software Engineer'))

# conn.commit()

# # Now fetch all rows and export
# os.makedirs('resumes', exist_ok=True)

# cursor.execute("SELECT id, full_name, email, phone_number, resume, job_position, applied_at FROM pravahapp_jobapplication")
# rows = cursor.fetchall()

# data_for_excel = []

# for row in rows:
#     id, full_name, email, phone, resume_blob, position, applied_at = row
#     filename = f"resume_{id}.pdf"
#     filepath = os.path.join('resumes', filename)

#     # Save the resume
#     with open(filepath, 'wb') as f:
#         f.write(resume_blob)

#     # Add row to export
#     data_for_excel.append({
#         'Full Name': full_name,
#         'Email': email,
#         'Phone': phone,
#         'Job Position': position,
#         'Applied At': applied_at,
#         'Resume File': filepath
#     })

# # Export to Excel
# df = pd.DataFrame(data_for_excel)
# df.to_excel('job_applications.xlsx', index=False)

# print("✔ Excel created and resumes saved.")

# # Cleanup
# cursor.close()
# conn.close()

# import mysql.connector
# import os
# import pandas as pd

# # Connect to MySQL
# conn = mysql.connector.connect(
#     host="localhost",
#     user="root",
#     password="sonali@12",  # Use your actual password
#     database="pravahdb"
# )

# cursor = conn.cursor()

# # Create folder to store resumes
# os.makedirs('resumes', exist_ok=True)

# # Fetch records
# cursor.execute("SELECT id, full_name, email, phone_number, resume, job_position, applied_at FROM pravahapp_jobapplication")
# rows = cursor.fetchall()

# data_for_excel = []

# for row in rows:
#     id, full_name, email, phone, resume_blob, position, applied_at = row

#     # Save the resume as a PDF file
#     resume_filename = f"resume_{id}_{full_name.replace(' ', '_')}.pdf"
#     resume_path = os.path.join("resumes", resume_filename)

#     with open(resume_path, 'wb') as f:
#         f.write(resume_blob)

#     # Add file path in the Excel row instead of BLOB
#     data_for_excel.append({
#         'Full Name': full_name,
#         'Email': email,
#         'Phone': phone,
#         'Job Position': position,
#         'Applied At': applied_at,
#         'Resume File': resume_path  # path instead of blob
#     })

# # Save Excel file
# df = pd.DataFrame(data_for_excel)
# df.to_excel('job_applications_with_resume_links.xlsx', index=False)

# print("✅ Excel created with resume links, and resumes saved in 'resumes/' folder.")

# import mysql.connector
# import os
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment

# # 1. Connect to your MySQL database
# conn = mysql.connector.connect(
#     host="localhost",
#     user="root",
#     password="sonali@12",  # Replace with your actual password
#     database="pravahdb"
# )

# cursor = conn.cursor()

# # 2. Create folder to save resumes
# os.makedirs("resumes", exist_ok=True)

# # 3. Query data from the job application table
# cursor.execute("""
#     SELECT id, full_name, email, phone_number, resume, job_position, applied_at
#     FROM pravahapp_jobapplication
# """)
# rows = cursor.fetchall()

# # 4. Process and save each resume as a PDF file
# excel_data = []

# for row in rows:
#     id, full_name, email, phone, resume_blob, job_position, applied_at = row

#     # Create filename
#     safe_name = full_name.replace(" ", "_")
#     pdf_filename = f"resume_{id}_{safe_name}.pdf"
#     pdf_path = os.path.abspath(os.path.join("resumes", pdf_filename))

#     # Save PDF to file
#     with open(pdf_path, "wb") as f:
#         f.write(resume_blob)

#     # Add data for Excel
#     excel_data.append({
#         "Full Name": full_name,
#         "Email": email,
#         "Phone Number": phone,
#         "Job Position": job_position,
#         "Applied At": applied_at,
#         "Resume File": f'=HYPERLINK("{pdf_path}", "Open Resume")'
#     })

# # 5. Save all data to Excel
# df = pd.DataFrame(excel_data)
# df.to_excel("job_applications_with_resumes.xlsx", index=False)


# wb = load_workbook("job_applications_with_resumes.xlsx")
# ws = wb.active

# for row in ws.iter_rows():
#     for cell in row:
#         cell.alignment = openpyxl.styles.Alignment(wrapText=True)
# wb.save("job_applications_with_resumes.xlsx")

# # Adjust column width in openpyxl
# ws.column_dimensions['F'].width = 50  # Adjust column F (the 'Resume File' column) width
# wb.save("job_applications_with_resumes.xlsx")

# # Cleanup

# cursor.close()
# conn.close()

# print("✅ Resumes saved in 'resumes' folder.")
# print("✅ Excel file created: job_applications_with_resumes.xlsx")


import mysql.connector
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 1. Connect to your MySQL database
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="sonali@12",  # Replace with your actual password
    database="pravahdb"
)

cursor = conn.cursor()

# 2. Create folder to save resumes
os.makedirs("resumes", exist_ok=True)

# 3. Query data from the job application table
cursor.execute("""
    SELECT id, full_name, email, phone_number, resume, job_position, applied_at
    FROM pravahapp_jobapplication
""")
rows = cursor.fetchall()

# 4. Process and save each resume as a PDF file
excel_data = []

for row in rows:
    id, full_name, email, phone, resume_blob, job_position, applied_at = row

    # Create filename
    safe_name = full_name.replace(" ", "_")
    pdf_filename = f"resume_{id}_{safe_name}.pdf"
    pdf_path = os.path.abspath(os.path.join("resumes", pdf_filename))

    # Save PDF to file
    with open(pdf_path, "wb") as f:
        f.write(resume_blob)

    # Add data for Excel
    excel_data.append({
        "Full Name": full_name,
        "Email": email,
        "Phone Number": phone,
        "Job Position": job_position,
        "Applied At": applied_at,
        "Resume File": f'=HYPERLINK("{pdf_path}", "Open Resume")'
    })

# 5. Save all data to Excel
df = pd.DataFrame(excel_data)
df.to_excel("job_applications_with_resumes.xlsx", index=False)

# 6. Open the workbook and adjust formatting
wb = load_workbook("job_applications_with_resumes.xlsx")
ws = wb.active

# Set text wrap for all cells and adjust column width
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrapText=True)

# Adjust column F width (Resume File)
ws.column_dimensions['F'].width = 80

# Save the final changes
wb.save("job_applications_with_resumes.xlsx")

# Cleanup
cursor.close()
conn.close()

print("✅ Resumes saved in 'resumes' folder.")
print("✅ Excel file created: job_applications_with_resumes.xlsx")
